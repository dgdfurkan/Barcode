"""Excel export fonksiyonları"""
import logging
from typing import List, Dict, Any, Optional
from pathlib import Path
from datetime import datetime
import io

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.formula.translate import Translator
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("⚠️ openpyxl modülü yüklü değil - Excel export çalışmayacak")


def create_stocks_excel(stocks: List[Dict[str, Any]], output_path: Optional[Path] = None) -> Optional[Path]:
    """Stok verilerini Excel dosyası olarak oluşturur
    
    Args:
        stocks: Stok verileri listesi
        output_path: Çıktı dosyası yolu (None ise otomatik oluşturulur)
        
    Returns:
        Oluşturulan Excel dosyasının yolu veya None (hata durumunda)
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("❌ openpyxl modülü yüklü değil - Excel export yapılamıyor")
        return None
    
    if not stocks:
        logger.warning("⚠️ Stok verisi boş - Excel dosyası oluşturulamıyor")
        return None
    
    try:
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Stoklar"
        
        # Başlık satırı
        headers = [
            "Ürün Adı",
            "Mevcut Stok",
            "Rezerve Stok",
            "Toplam",
            "Raf Etiketi",
            "Görsel Linki",
            "Görsel"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
        
        # Stok verilerini ekle
        row_num = 2
        for stock in stocks:
            # Ürün bilgilerini çıkar
            product = stock.get('product', {})
            if isinstance(product, dict):
                # Ürün adı
                name = product.get('fullName', '') or product.get('name', {})
                if isinstance(name, dict):
                    name = name.get('tr', 'Bilinmeyen Ürün')
                if not name:
                    name = 'Bilinmeyen Ürün'
                
                # Barkod/raf etiketi
                barcodes = product.get('barcodes', [])
                if not barcodes and isinstance(product.get('packagingInfo'), dict):
                    packaging_info = product.get('packagingInfo', {})
                    type_1 = packaging_info.get('1', {})
                    if isinstance(type_1, dict) and 'barcodes' in type_1:
                        barcodes_list = type_1.get('barcodes', [])
                        if isinstance(barcodes_list, list) and len(barcodes_list) > 0:
                            barcodes = [barcodes_list[0]]
                
                barcode = barcodes[0] if barcodes else "Bulunamadı"
                
                # Görsel linki
                pic_url = product.get('picURL') or product.get('imageUrl') or product.get('image')
                if not pic_url and isinstance(product.get('images'), list) and len(product.get('images', [])) > 0:
                    pic_url = product.get('images')[0]
                
                # Görsel linkini tam URL'ye çevir (eğer relative ise)
                if pic_url and not pic_url.startswith('http'):
                    # Getir CDN URL'si
                    pic_url = f"https://cdn.getir.com{pic_url}" if pic_url.startswith('/') else f"https://cdn.getir.com/{pic_url}"
            else:
                name = 'Bilinmeyen Ürün'
                barcode = "Bulunamadı"
                pic_url = None
            
            # Stok bilgileri
            available = stock.get('available', 0)
            reserve = stock.get('reserve', 0)
            total = available + reserve
            
            # Sayımda durumu
            if available == -1:
                available_display = "Sayımda"
                total_display = "Sayımda"
            else:
                available_display = available
                total_display = total
            
            # Satır verilerini ekle
            ws.cell(row=row_num, column=1, value=name)
            ws.cell(row=row_num, column=2, value=available_display)
            ws.cell(row=row_num, column=3, value=reserve)
            ws.cell(row=row_num, column=4, value=total_display)
            ws.cell(row=row_num, column=5, value=barcode)
            ws.cell(row=row_num, column=6, value=pic_url if pic_url else "")
            
            # Görsel için IMAGE() formülü (Excel'de görsel göstermek için)
            if pic_url:
                # IMAGE() formülü: =IMAGE(url, [alt_text], [sizing], [height], [width])
                # sizing: 0=fit cell, 1=original size, 2=custom size
                # Excel'de IMAGE() formülü için URL'yi escape et
                escaped_url = pic_url.replace('"', '""')  # Excel'de çift tırnak escape edilir
                image_formula = f'=IMAGE("{escaped_url}", 1, 1, 100, 100)'
                ws.cell(row=row_num, column=7, value=image_formula)
            else:
                ws.cell(row=row_num, column=7, value="")
            
            row_num += 1
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 40  # Ürün Adı
        ws.column_dimensions['B'].width = 15  # Mevcut Stok
        ws.column_dimensions['C'].width = 15  # Rezerve Stok
        ws.column_dimensions['D'].width = 15  # Toplam
        ws.column_dimensions['E'].width = 20  # Raf Etiketi
        ws.column_dimensions['F'].width = 50  # Görsel Linki
        ws.column_dimensions['G'].width = 20  # Görsel
        
        # Dosya yolunu belirle
        if output_path is None:
            from src.utils import get_project_root
            project_root = get_project_root()
            data_dir = project_root / "data"
            data_dir.mkdir(exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = data_dir / f"stoklar_{timestamp}.xlsx"
        
        # Dosyayı kaydet
        wb.save(output_path)
        logger.info(f"✅ Excel dosyası oluşturuldu: {output_path} ({len(stocks)} ürün)")
        
        return output_path
        
    except Exception as e:
        logger.error(f"❌ Excel dosyası oluşturulurken hata: {e}", exc_info=True)
        return None

